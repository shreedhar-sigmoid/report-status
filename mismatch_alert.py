from smtplib import SMTP_SSL
from os import getenv, remove
from openpyxl import Workbook
from dotenv import load_dotenv
from email.mime.text import MIMEText
from ssl import create_default_context
from mongo_collection import get_mongo_coll
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

load_dotenv()

SMTP_PORT = getenv("SMTP_PORT")
SMTP_SERVER = getenv("SMTP_SERVER")
SENDER_EMAIL = getenv("SENDER_EMAIL")
RECIVER_EMAIL = getenv("RECIVER_EMAIL")
EMAIL_PASSWORD = getenv("EMAIL_PASSWORD")


def send_mismatch_alert(all_data_list, previous_day_epoch, today_epoch, yesterday_date):
    """
    send_mismatch_alert function sends the mail for mismatch report counts

    previous_day_epoch : Epoch time for state date
    today_epoch:       : Epoch time for end date
    yesterday_date     : Yesterday date in MM-DD-YYYY format
    report_coll, uns_coll, reportemail_coll, reportalert_coll
    """
    report_coll, uns_coll, reportemail_coll, reportalert_coll = get_mongo_coll()
    mismatch_message = MIMEMultipart("alternative")
    mismatch_message["Subject"] = "[ALERT] Mismatch Reports[{yesterday_date}]".format(
        yesterday_date=yesterday_date
    )
    mismatch_message["From"] = SENDER_EMAIL
    mismatch_message["To"] = RECIVER_EMAIL
    mismatch_count = 0
    mismatch_row = ""
    wb = Workbook()
    for reports_list in all_data_list:
        # handling exception for monthly report
        if reports_list[2] != []:
            for report in reports_list[2]:
                estimated_reports_number = report[2]
                count_of_all_states = sum(report[3:])

                if estimated_reports_number != count_of_all_states:

                    mismatch_count += 1
                    org = report[0]
                    view = report[1]
                    triggerType = reports_list[0]
                    triggerDay = reports_list[1]
                    report_email_list = []
                    active_users_emails = []
                    estimated_reports_names = []

                    active_reports = report_coll.find(
                                                {
                                                    "reportStatus": "active",
                                                    "triggerType": triggerType,
                                                    "triggerDay": triggerDay,
                                                    "orgViewReq.organization": org,
                                                    "orgViewReq.view": view,
                                                    "createdOn": {"$lt": previous_day_epoch},
                                                },
                                                {"loginInfo.providerKey": 1, "_id": 1},
                                            )
                    active_users = uns_coll.find(
                                            {
                                                "activeStatus": True,
                                                "orgInfoList.name": org,
                                                "orgInfoList.viewInfoList.name": view,
                                            },
                                            {"email": 1},
                                        )
                    sent_reports = reportemail_coll.find(
                                                    {
                                                        "emailSentTime": {
                                                            "$gt": previous_day_epoch,
                                                            "$lt": today_epoch,
                                                        },
                                                        "emailStatus": "c",
                                                        "triggerType": triggerType,
                                                        "orgViewReq.organization": org,
                                                        "orgViewReq.view": view,
                                                    },
                                                    {"_id": 1},
                                                )
                    pending_reports = reportalert_coll.find(
                                                        {
                                                            "executionOn": {
                                                                "$gt": previous_day_epoch,
                                                                "$lt": today_epoch,
                                                            },
                                                            "triggerStatus": "p",
                                                            "triggerType": triggerType,
                                                            "orgViewReq.organization": org,
                                                            "orgViewReq.view": view,
                                                        },
                                                        {"_id": 1},
                                                    )
                    inprogess_reports = reportalert_coll.find(
                                                        {
                                                            "executionOn": {
                                                                "$gt": previous_day_epoch,
                                                                "$lt": today_epoch,
                                                            },
                                                            "triggerStatus": "i",
                                                            "triggerType": triggerType,
                                                            "orgViewReq.organization": org,
                                                            "orgViewReq.view": view,
                                                        },
                                                        {"_id": 1},
                                                    )
                    no_data_reports = reportalert_coll.find(
                                                        {
                                                            "executionOn": {
                                                                "$gt": previous_day_epoch,
                                                                "$lt": today_epoch,
                                                            },
                                                            "triggerStatus": "ND",
                                                            "triggerType": triggerType,
                                                            "orgViewReq.organization": org,
                                                            "orgViewReq.view": view,
                                                        },
                                                        {"_id": 1},
                                                    )
                    failed_reports = reportalert_coll.find(
                                                        {
                                                            "executionOn": {
                                                                "$gt": previous_day_epoch,
                                                                "$lt": today_epoch,
                                                            },
                                                            "triggerStatus": "f",
                                                            "triggerType": triggerType,
                                                            "orgViewReq.organization": org,
                                                            "orgViewReq.view": view,
                                                        },
                                                        {"_id": 1},
                                                    )

                    for report in active_reports:
                        report_email_list.append(
                            [report["loginInfo"]["providerKey"], report["_id"]]
                        )

                    for email in active_users:
                        active_users_emails.append(email["email"])

                    for report in report_email_list:  # Selecting only valid active reports count
                        if report[0] in active_users_emails:
                            estimated_reports_names.append(report[1])

                    def return_list(result_list):  # Extracting the IDs from mongo collection
                        holder_list = []
                        for item in result_list:
                            holder_list.append(item["_id"]) 

                        return holder_list

                    report_names_list = [
                        sent_reports,
                        pending_reports,
                        inprogess_reports,
                        no_data_reports,
                        failed_reports,
                    ]
                    mismatch_reports_names_list = [estimated_reports_names] # Initialize list with estimated_reports_names

                    for x in report_names_list:
                        report_list = return_list(x)
                        mismatch_reports_names_list.append(report_list)

                    
                    name = org + "_" + view + ".xlsx"
                    mismatch = wb.create_sheet(name) # Creating an xlsx sheet with org and view name

                    xlsx_header_list = [
                        "Estimated_reports",
                        "Total_sent_reports",
                        "Total_pending_reports",
                        "Total_inprogess_reports",
                        "Total_No_date_reports",
                        "Total_failed_reports",
                    ]

                    
                    for index, item in enumerate(xlsx_header_list):  # Adding headers to sheet
                        mismatch.cell(1, index + 1).value = item

                    
                    for index1, row in enumerate(mismatch_reports_names_list): # Adding report IDs to sheet
                        for index2, col in enumerate(row):
                            mismatch.cell(row=index2 + 3, column=index1 + 1).value = col

                    mismatch_row += (
                        """ <tr style='background-color: #E1E5EA'>
                                <td> """ + org + """ </td>
                                <td> """ + view + """ </td>
                                <td> """ + triggerType + """ </td>
                                <td> """ + str(estimated_reports_number) + """</td>
                                <td> """ + str(count_of_all_states) + """ </td>
                            </tr>"""
                    )
    html = """\
        <html>
            <body>
                <h2 style="color:red;"><span style="font-size:20px;">❌</span> Mismatch found in the followings </h2>
                <h3>Date: {yesterday_date}</h3>
                <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                    <tr style='background-color: #203239;color:#F7F7F7'>
                        <th> Organization </th>
                        <th> View </th>
                        <th> Trigger type </th>
                        <th> Estimated count </th>
                        <th> All states count </th>
                    </tr>
                    {Mismatch_row}
                </table>
                <h4>Please find the reports IDs in the attatchments bellow.</h4>
            </body>
        </html>""".format(
        Mismatch_row=mismatch_row, yesterday_date=yesterday_date
    )

    if mismatch_count > 0:
        wb.remove(wb["Sheet"])  # removing default sheet from xlsx file
        xl_name = "mismatch-" + yesterday_date + "-.xlsx"
        wb.save(xl_name)
        attachment = open(xl_name, "rb")
        part = MIMEApplication(attachment.read(), _subtype="xlsx")
        part.add_header("Content-Disposition", "attachment", filename=xl_name)
        mismatch_message.attach(part)
        part1 = MIMEText(html, "html")
        mismatch_message.attach(part1)
        try:
            context = create_default_context()
            with SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as server:
                server.login(SENDER_EMAIL, EMAIL_PASSWORD)
                server.sendmail(
                    SENDER_EMAIL, RECIVER_EMAIL, mismatch_message.as_string()
                )
                remove(xl_name)  # Deleting the created xlsx file from local storage
        except:
            print("Something went wrong")
