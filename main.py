from urllib import parse
from requests import post
from smtplib import SMTP_SSL
from json import dumps, loads
from openpyxl import Workbook
from dotenv import load_dotenv
from pymongo import MongoClient
from email.mime.text import MIMEText
from os import getenv , remove , path
from ssl import create_default_context
from email.mime.multipart import MIMEMultipart
from datetime import date ,timedelta, datetime
from email.mime.application import MIMEApplication
from pickle import load as pickle_load, dump as pickle_dump


#loading all envirment varibles
load_dotenv()

ORG_LIST = loads(getenv('ORG_LIST'))
EXCEPTION_LIST = loads(getenv('EXCEPTION_LIST'))
SIGVIEW_EMAIL= getenv('SIGVIEW_EMAIL')
SIGVIEW_PASSWORD= getenv('SIGVIEW_PASSWORD')
GET_TOKEN_URL= getenv('GET_TOKEN_URL')
GET_DATE_RANGE_URL= getenv('GET_DATE_RANGE_URL')

#Mongo env varibables
MONGO_USERNAME= getenv('MONGO_USERNAME')
MONOGO_PASSWORD= getenv('MONOGO_PASSWORD')
MONGO_HOST= getenv('MONGO_HOST')
MONGO_PORT= getenv('MONGO_PORT')
MONDODB= getenv('MONDODB')

#email variables
SMTP_PORT= getenv('SMTP_PORT')
SMTP_SERVER= getenv('SMTP_SERVER')
SENDER_EMAIL= getenv('SENDER_EMAIL')
RECIVER_EMAIL= getenv('RECIVER_EMAIL')
EMAIL_PASSWORD= getenv('EMAIL_PASSWORD')


#getting epoch time 
today_epoch = str((int(date.today().strftime("%s")) *1000)+19800000)
date_lag_epoch = (int(date.today().strftime("%s")) *1000)+27000000
monthly_epoch = int(date.today().replace(day=2).strftime("%s")) *1000 + 19800000

#funtion to get x-authtoken
def get_token():

    headers = {"content-type": "application/json;charset=UTF-8"}
    data = {"email": SIGVIEW_EMAIL, "password": SIGVIEW_PASSWORD , "rememberMe": True}
    data = dumps(data)
    response = post(url=GET_TOKEN_URL, headers=headers, data=data)
    response = response.content.decode("utf-8")
    response = loads(response)
    token = response["token"]

    return token


#funtion to get data lag 
def date_range(token):

    data_lag = []
    headers = {
        "X-Auth-Token": token,
        "Content-Type": "application/json;charset=UTF-8"
    }

    for org, view in ORG_LIST:

        org_view = org+"_"+view
        if org_view not in EXCEPTION_LIST:
            data = {"organization":org,"view":view}
            data = dumps(data)
            try:
                response = post(url=GET_DATE_RANGE_URL, headers=headers, data=data)
                response = response.content.decode("utf-8")
                response = loads(response)
                data_lag.append(int(response["result"]["endDate"]))
            except:
                print(f'You dont have access to {org} and {view}')
                return None
        
    return data_lag

def report_tracker(today_epoch,previous_day_epoch,day_of_week,yesterday_date,last_month):

    username = parse.quote_plus(MONGO_USERNAME)
    password = parse.quote_plus(MONOGO_PASSWORD)
   
    tracker_message = MIMEMultipart("alternative")
    tracker_message["Subject"] = "Reports status[{yesterday_date}]".format(yesterday_date=yesterday_date)
    tracker_message["From"] = SENDER_EMAIL
    tracker_message["To"] = RECIVER_EMAIL

    #Mongo connection
    CONNECTION_STRING = f'mongodb://{username}:{password}@{MONGO_HOST}:{MONGO_PORT}/{MONDODB}?authMechanism=SCRAM-SHA-1'
    mongo_database = MongoClient(CONNECTION_STRING).get_database('sigview2')
    report_coll = mongo_database.get_collection('report')
    uns_coll = mongo_database.get_collection('users_new_sample')
    reportemail_coll = mongo_database.get_collection('reportemail')
    reportalert_coll = mongo_database.get_collection('reportalert')

    def get_reports(ORG_LIST,triggerType,triggerDay):

        data = []

        for org,view in ORG_LIST:

            report_email_list = []
            active_users_emails = []
            estimated_reports = 0

            active_reports = report_coll.find({"reportStatus":"active","triggerType" : triggerType,"triggerDay" : triggerDay,"orgViewReq.organization" : org,"orgViewReq.view" : view,"createdOn" : {"$lt":previous_day_epoch}},{"loginInfo.providerKey":1})
            active_users = uns_coll.find({"activeStatus":True,"orgInfoList.name":org,"orgInfoList.viewInfoList.name":view},{"email":1})
            total_sent_reports = reportemail_coll.count({"emailSentTime":{"$gt":previous_day_epoch,"$lt":today_epoch},"emailStatus" : "c","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view})
            total_pending_reports = reportalert_coll.count({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "p","triggerType":triggerType,"orgViewReq.organization":org, "orgViewReq.view":view})
            total_inprogess_reports = reportalert_coll.count({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "i","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view})
            total_No_date_reports = reportalert_coll.count({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "ND","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view})
            total_failed_reports = reportalert_coll.count({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "f","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view})
                       
            for report in active_reports:
                report_email_list.append(report["loginInfo"]["providerKey"])
            
            for email in active_users:
                active_users_emails.append(email["email"])
            
            for count in report_email_list:
                if count in active_users_emails:
                    estimated_reports +=1
            
            result=[org, view, estimated_reports, total_sent_reports, total_pending_reports, total_inprogess_reports, total_No_date_reports, total_failed_reports]
            data.append(result)

        return data
    
    daily_reports = get_reports(ORG_LIST,"daily","everyday")
    weekly_reports = get_reports(ORG_LIST,"weekly",day_of_week)
    monthly_reports = []

    def table_row(rows):
        t=""
        for row in rows:
            if row[2] != sum(row[3:]):
                t+="""<tr style='background-color: #FFBBCC;'>
                    <td>"""+ row[0] +"""</td>
                    <td>"""+ row[1] +"""</td>
                    <td>"""+ str(row[2]) +"""</td>
                    <td>"""+ str(row[3]) +"""</td>
                    <td>"""+ str(row[4]) +"""</td>
                    <td>"""+ str(row[5]) +"""</td>
                    <td>"""+ str(row[6]) +"""</td>
                    <td>"""+ str(row[7]) +"""</td>
                </tr>"""
            else:
                t+="""<tr style='background-color: #E1E5EA'>
                <td>"""+ row[0] +"""</td>
                <td>"""+ row[1] +"""</td>
                <td>"""+ str(row[2]) +"""</td>
                <td>"""+ str(row[3]) +"""</td>
                <td>"""+ str(row[4]) +"""</td>
                <td>"""+ str(row[5]) +"""</td>
                <td>"""+ str(row[6]) +"""</td>
                <td>"""+ str(row[7]) +"""</td>
            </tr>"""
        return t

    monthly_table=""

    if today_epoch == monthly_epoch:
        monthly_reports = get_reports(ORG_LIST,"monthly","1st of Every Month")
        monthly_table = """
                <h2>Monthly report details:</h2>
                <h3>Month: {last_month} </h3>
                <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                    <tr style='background-color: #203239;color:#F7F7F7'>
                        <th>Organization</th>
                        <th>View</th>
                        <th>Monthly_Estimated_Mail</th>
                        <th>Total_Sent_email</th>
                        <th>Pending_Status</th>
                        <th>Inprocess_Status</th>
                        <th>No_Data</th>
                        <th>Failed_Status</th> 
                    </tr>""".format(last_month=last_month)
        monthly_table+=table_row(monthly_reports)
        monthly_table+="</table>"

    html = """\
    <html>
        <body>
            <h2>Daily report details:</h2>
            <h3>Date: {yesterday_date} </h3>
            <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                <tr style='background-color: #203239;color:#F7F7F7'>
                    <th>Organization</th>
                    <th>View</th>
                    <th>Daily_Estimated_Mails</th>
                    <th>Total_Sent_Mails</th>
                    <th>Pending_Status</th>
                    <th>Inprocess_Status</th>
                    <th>No_Data</th>
                    <th>Failed_Status</th> 
                </tr>
                {Daily_Row}
            </table>
            <h2>Weekly report details for {day_of_week}:</h2>
            <h3>Date: {yesterday_date} </h3>
            <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                <tr style='background-color: #203239;color:#F7F7F7'>
                    <th>Organization</th>
                    <th>View</th>
                    <th>Weekly_Estimated_Mail</th>
                    <th>Total_Sent_email</th>
                    <th>Pending_Status</th>
                    <th>Inprocess_Status</th>
                    <th>No_Data</th>
                    <th>Failed_Status</th> 
                </tr>
                {Weekly_Row}
            </table>
                {Monthly_Table}
        </body>
    </html>
    """.format(Daily_Row = table_row(daily_reports),yesterday_date=yesterday_date,Weekly_Row=table_row(weekly_reports),day_of_week= day_of_week,Monthly_Table=monthly_table)

    part1 = MIMEText(html,"html")
    tracker_message.attach(part1)

    context = create_default_context()
    with SMTP_SSL(SMTP_SERVER,SMTP_PORT, context=context) as server:
        server.login(SENDER_EMAIL,EMAIL_PASSWORD)
        server.sendmail(SENDER_EMAIL,RECIVER_EMAIL,tracker_message.as_string())


    #Funtion to send mismatch reports
    def mismatch_alert(all_data_list):

        mismatch_message = MIMEMultipart("alternative")
        mismatch_message["Subject"] = "[ALERT] Mismatch Reports[{yesterday_date}]".format(yesterday_date=yesterday_date)
        mismatch_message["From"] = SENDER_EMAIL
        mismatch_message["To"] = RECIVER_EMAIL
        mismatch_count = 0
        mismatch_row =""
        wb = Workbook()   
        for reports_list in all_data_list:
            #handling exception for monthly report
            if reports_list[2] != []:
                for report in reports_list[2]:
                    estimated_reports_number = report[2]
                    count_of_all_states = sum(report[3:])
                    
                    if estimated_reports_number != count_of_all_states:
                        
                        mismatch_count +=1

                        org = report[0]
                        view = report[1]
                        triggerType = reports_list[0]
                        triggerDay = reports_list[1]
                        report_email_list = []
                        active_users_emails = []
                        estimated_reports_names = []

                        active_reports = report_coll.find({"reportStatus":"active","triggerType" : triggerType,"triggerDay" : triggerDay,"orgViewReq.organization" : org,"orgViewReq.view" : view,"createdOn" : {"$lt":previous_day_epoch}},{"loginInfo.providerKey":1,"_id":1})
                        active_users = uns_coll.find({"activeStatus":True,"orgInfoList.name":org,"orgInfoList.viewInfoList.name":view},{"email":1})
                        sent_reports = reportemail_coll.find({"emailSentTime":{"$gt":previous_day_epoch,"$lt":today_epoch},"emailStatus" : "c","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view},{"_id":1})
                        pending_reports = reportalert_coll.find({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "p","triggerType":triggerType,"orgViewReq.organization":org, "orgViewReq.view":view},{"_id":1})
                        inprogess_reports = reportalert_coll.find({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "i","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view},{"_id":1})
                        no_data_reports = reportalert_coll.find({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "ND","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view},{"_id":1})
                        failed_reports = reportalert_coll.find({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "f","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view},{"_id":1})
                        
                        for report in active_reports:
                            report_email_list.append([report["loginInfo"]["providerKey"],report["_id"]])
                        
                        for email in active_users:
                            active_users_emails.append(email["email"])
                        
                        for report in report_email_list:
                            if report[0] in active_users_emails:
                                estimated_reports_names.append(report[1])
                            
                        def return_list(result_list):

                            holder_list = []
                            for item in result_list:
                                holder_list.append(item["_id"])

                            return holder_list
                        report_names_list = [ sent_reports, pending_reports, inprogess_reports, no_data_reports, failed_reports ]
                        mismatch_reports_names_list = [estimated_reports_names]
                        
                        for x in report_names_list:
                            report_list = return_list(x)
                            mismatch_reports_names_list.append(report_list)
                        
                        #creating an xl sheet 
                        name = org+"_"+view+".xlsx"
                        mismatch = wb.create_sheet(name) 
                        
                        xlsx_header_list = ["Estimated_reports", "Total_sent_reports", "Total_pending_reports", "Total_inprogess_reports", "Total_No_date_reports", "Total_failed_reports"]
                        
                        #adding headers to sheet
                        for index, item in enumerate(xlsx_header_list):
                            mismatch.cell(1,index+1).value = item
                        
                        #adding report names to sheet
                        for index1,row in enumerate(mismatch_reports_names_list):
                            for index2,col in enumerate(row):
                                mismatch.cell(row=index2+3,column=index1+1).value = col

                        #removing default sheet
                        mismatch_row += """ <tr style='background-color: #E1E5EA'>
                                                <td> """+ org +""" </td>
                                                <td> """+ view +""" </td>
                                                <td> """+ triggerType +""" </td>
                                                <td> """+ str(estimated_reports_number) +"""</td>
                                                <td> """+ str(count_of_all_states) +""" </td>
                                            </tr>"""
        html = """\
            <html>
                <body>
                    <h2 style="color:red;"><span style="font-size:20px;">❌</span> Mismatch found in the followings </h2>
                    <h3>Date: {yesterday_date}</h3>
                    <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                        <tr style='background-color: #203239;color:#F7F7F7'>
                            <th> Organization </th>
                            <th> View </th>
                            <th> Trigger type </th>
                            <th> Estimated count </th>
                            <th> All states count </th>
                        </tr>
                        {Mismatch_row}
                    </table>
                 <h4>Please find the reports IDs in the attatchments bellow.</h4>
                </body>
            </html>""".format(Mismatch_row=mismatch_row,yesterday_date=yesterday_date)
        
        if mismatch_count > 0 :
            wb.remove(wb["Sheet"])
            xl_name = "mismatch-"+ yesterday_date +"-.xlsx"
            wb.save(xl_name)
            attachment = open(xl_name, 'rb')
            part = MIMEApplication(attachment.read(), _subtype='xlsx')
            part.add_header('Content-Disposition', 'attachment', filename=xl_name)
            mismatch_message.attach(part)
            part1 = MIMEText(html,"html")
            mismatch_message.attach(part1)
            try:
                context = create_default_context()
                with SMTP_SSL(SMTP_SERVER,SMTP_PORT, context=context) as server:
                    server.login(SENDER_EMAIL,EMAIL_PASSWORD)
                    server.sendmail(SENDER_EMAIL,RECIVER_EMAIL,mismatch_message.as_string())
                    remove(xl_name)
            except:
                print("Something went wrong")


    def failed_alert(all_data_list):
    
        failed_message = MIMEMultipart("alternative")
        failed_message["Subject"] = "[ALERT] Failed Reports [{yesterday_date}]".format(yesterday_date=yesterday_date)
        failed_message["From"] = SENDER_EMAIL
        failed_message["To"] = RECIVER_EMAIL
        fail_count = 0
        failed_row =""
        wb = Workbook()   
        for reports_list in all_data_list:

            #handling exception for monthly report
            if reports_list[2] != []:
                for report in reports_list[2]:
                    if report[7] > 0:

                        fail_count += 1

                        org = report[0]
                        view = report[1]
                        triggerType = reports_list[0]
                        failed_reports_number = report[7]

                        failed_reports = reportalert_coll.find({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "f","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view},{"_id":1})
                        failed_reports_names =[]
                            
                        for item in failed_reports:
                            failed_reports_names.append(item["_id"])
                    
                        #creating an xl sheet 
                        name = org+"_"+view+".xlsx"
                        failed = wb.create_sheet(name) 
                        failed.cell(1,1).value = "Failed report names"

                        #adding failed reportnames to sheet
                        for index, item in enumerate(failed_reports_names):
                            failed.cell(index+2,1).value = item
                        
                        failed_row += """ <tr style='background-color: #E1E5EA'>
                                                <td> """+ org +""" </td>
                                                <td> """+ view +""" </td>
                                                <td> """+ triggerType +""" </td>
                                                <td> """+ str(failed_reports_number) +"""</td>
                                            </tr>"""
        html = """\
                    <html>
                        <body>
                            <h2 style="color:red;"><span style="font-size:20px;">❌</span> Failed report found in the followings </h2>
                            <h3>Date: {yesterday_date}</h3>
                            <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                                <tr style='background-color: #203239;color:#F7F7F7'>
                                    <th> Organization </th>
                                    <th> View </th>
                                    <th> Trigger type </th>
                                    <th> Failed state count </th>
                                </tr>
                                {failed_row}
                            </table>
                            <h4>Please find the reports IDs in the attatchments bellow.</h4>
                        </body>
                    </html>""".format(failed_row=failed_row,yesterday_date=yesterday_date)

        if fail_count > 0:
            #removing default sheet
            wb.remove(wb["Sheet"])
            xl_name = "failed-" + yesterday_date +"-.xlsx"
            wb.save(xl_name)
            attachment = open(xl_name, 'rb')
            part = MIMEApplication(attachment.read(), _subtype='xlsx')
            part.add_header('Content-Disposition', 'attachment', filename=xl_name)
            failed_message.attach(part)
            part1 = MIMEText(html,"html")
            failed_message.attach(part1)
            try:
                context = create_default_context()
                with SMTP_SSL(SMTP_SERVER,SMTP_PORT, context=context) as server:
                    server.login(SENDER_EMAIL,EMAIL_PASSWORD)
                    server.sendmail(SENDER_EMAIL,RECIVER_EMAIL,failed_message.as_string())
                    remove(xl_name)
            except:
                print("Something went wrong")


    #sending mails for mismatch and failed reports
    all_data_list = [["daily","everyday",daily_reports ],["weekly",day_of_week, weekly_reports ],["montly","1st of Every Month",monthly_reports ]]
    mismatch_alert(all_data_list)
    failed_alert(all_data_list)
    


token = get_token()
data_lag=date_range(token)


if data_lag != None:
    sort_data_lag = data_lag[:]
    sort_data_lag.sort()
    if 1==1:
    #checking if there any data lag less than 2nd hour of today(UTC)
    #if sort_data_lag[0] >= data_lag_epoch:
        if not path.exists("sendtime.pkl"):
            today_epoch = str((int(date.today().strftime("%s")) *1000)+19800000)
            previous_day_epoch = str((int((date.today() - timedelta(days=1)).strftime("%s")) *1000)+19800000)   
            day_of_week = datetime.fromtimestamp(int(previous_day_epoch)/1000).strftime("%A").lower()
            yesterday_date = (date.today() - timedelta(days = 1)).strftime("%d-%m-%Y")
            last_month = (date.today().replace(day = 1) - timedelta(days=1)).strftime("%B")  
            report_tracker(today_epoch,previous_day_epoch,day_of_week,yesterday_date,last_month)
            email_sent_time = today_epoch
            with open("sendtime.pkl", "wb") as file:
                pickle_dump(email_sent_time, file)

        else:
            email_sent_time = pickle_load(open("sendtime.pkl","rb"))
            if int(email_sent_time) < int(today_epoch):
                days = int((int(today_epoch) - int(email_sent_time))/86400000)
                for day in range(1,days+1):
                    today_epoch = str((int((date.today() - timedelta(days=day-1)).strftime("%s")) *1000)+19800000) 
                    previous_day_epoch = str((int((date.today() - timedelta(days=day)).strftime("%s")) *1000)+19800000) 
                    day_of_week = datetime.fromtimestamp(int(previous_day_epoch)/1000).strftime("%A").lower()
                    yesterday_date = (date.today() - timedelta(days = day)).strftime("%d-%m-%Y")
                    last_month = ((date.today() - timedelta(days=day+1)).replace(day = 1) - timedelta(days=1)).strftime("%B")
                    report_tracker(today_epoch,previous_day_epoch,day_of_week,yesterday_date,last_month)
                email_sent_time = str((int(date.today().strftime("%s")) *1000)+19800000)
                with open("sendtime.pkl", "wb") as file:
                    pickle_dump(email_sent_time, file)
