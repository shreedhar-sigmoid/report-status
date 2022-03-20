from smtplib import SMTP_SSL
from openpyxl import Workbook
from os import getenv ,remove 
from dotenv import load_dotenv
from email.mime.text import MIMEText
from ssl import create_default_context
from mongo_collection import get_mongo_coll
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


load_dotenv()

#email variables
SMTP_PORT= getenv('SMTP_PORT')
SMTP_SERVER= getenv('SMTP_SERVER')
SENDER_EMAIL= getenv('SENDER_EMAIL')
RECIVER_EMAIL= getenv('RECIVER_EMAIL')
EMAIL_PASSWORD= getenv('EMAIL_PASSWORD')



def send_failed_alert(all_data_list, previous_day_epoch, today_epoch,yesterday_date):
    """
    send_failed_alert function sends the mail for failed reports

    all_data_list      : count jobs state in the report, reportalert and reportemail as a list
    previous_day_epoch : epoch time for the state date
    today_epoch:       : epoch time for the end date
    yesterday_date     : yesterday date in MM-DD-YYYY format
    """

    reportalert_coll = get_mongo_coll()[3]

    failed_message = MIMEMultipart("alternative")
    failed_message["Subject"] = "[ALERT] Failed Reports [{yesterday_date}]".format(yesterday_date=yesterday_date)
    failed_message["From"] = SENDER_EMAIL
    failed_message["To"] = RECIVER_EMAIL
    fail_count = 0
    failed_row =""
    wb = Workbook()   
    for reports_list in all_data_list:

        #handling exception for monthly report
        if reports_list[2] != []:
            for report in reports_list[2]:
                if report[7] > 0:

                    fail_count += 1

                    org = report[0]
                    view = report[1]
                    triggerType = reports_list[0]
                    failed_reports_number = report[7]

                    failed_reports = reportalert_coll.find({"executionOn":{"$gt":previous_day_epoch,"$lt":today_epoch},"triggerStatus" : "f","triggerType":triggerType,"orgViewReq.organization":org,"orgViewReq.view":view},{"_id":1})
                    failed_reports_names =[]
                        
                    for item in failed_reports:
                        failed_reports_names.append(item["_id"])
                
                    #creating an xl sheet 
                    name = org+"_"+view+".xlsx"
                    failed = wb.create_sheet(name) 
                    failed.cell(1,1).value = "Failed report names"

                    #adding failed reportnames to sheet
                    for index, item in enumerate(failed_reports_names):
                        failed.cell(index+2,1).value = item
                    
                    failed_row += """ <tr style='background-color: #E1E5EA'>
                                            <td> """+ org +""" </td>
                                            <td> """+ view +""" </td>
                                            <td> """+ triggerType +""" </td>
                                            <td> """+ str(failed_reports_number) +"""</td>
                                        </tr>"""
    html = """\
                <html>
                    <body>
                        <h2 style="color:red;"><span style="font-size:20px;">❌</span> Failed report found in the followings </h2>
                        <h3>Date: {yesterday_date}</h3>
                        <table border='1' style='border-collapse:collapse;text-align: center; vertical-align: middle;'>
                            <tr style='background-color: #203239;color:#F7F7F7'>
                                <th> Organization </th>
                                <th> View </th>
                                <th> Trigger type </th>
                                <th> Failed state count </th>
                            </tr>
                            {failed_row}
                        </table>
                        <h4>Please find the reports IDs in the attatchments bellow.</h4>
                    </body>
                </html>""".format(failed_row=failed_row,yesterday_date=yesterday_date)

    if fail_count > 0:
        #removing default sheet
        wb.remove(wb["Sheet"])
        xl_name = "failed-" + yesterday_date +"-.xlsx"
        wb.save(xl_name)
        attachment = open(xl_name, 'rb')
        part = MIMEApplication(attachment.read(), _subtype='xlsx')
        part.add_header('Content-Disposition', 'attachment', filename=xl_name)
        failed_message.attach(part)
        part1 = MIMEText(html,"html")
        failed_message.attach(part1)
        try:
            context = create_default_context()
            with SMTP_SSL(SMTP_SERVER,SMTP_PORT, context=context) as server:
                server.login(SENDER_EMAIL,EMAIL_PASSWORD)
                server.sendmail(SENDER_EMAIL,RECIVER_EMAIL,failed_message.as_string())
                remove(xl_name)
        except:
            print("Something went wrong")